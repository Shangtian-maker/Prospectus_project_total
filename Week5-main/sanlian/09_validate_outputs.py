#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
09_validate_outputs.py

对以下两份工作簿进行 Cross-check 和 Schema 校验：

1. sanlian_gold.xlsx
   - subscription_flow_gold
   - share_transfer_flow_gold
   - equity_snapshot_gold

2. sanlian.xlsx
   - subscription_flow
   - share_transfer_flow
   - equity_snapshot

改进点：
- 固定识别上述六个工作表；
- 对“三联锻造”“三联有限”“芜湖三联锻造股份有限公司”等名称统一处理；
- 候选配对键不再包含“公司”，避免同一发行人的名称差异导致全部失配；
- Excel 日期对象、Excel 日期序列号、中文年月、ISO 日期统一按年月比较；
- t0|报告期初、t0，以及 t1 的不同描述按同一时点编号比较；
- 事件类型、价格、页码和原文证据属于软字段，差异不阻断 PASS；
- 抽取表比 Gold 多出的记录只记 WARNING，不阻断 PASS；
- Schema 只对抽取结果执行；结构错误为 ERROR，证据与闭环问题为 WARNING。

可信 PASS 条件：
- Gold 中每条记录均找到对应抽取记录；
- 核心业务字段没有差异；
- 抽取表没有 Schema ERROR。

输出：
- validation_results.xlsx
- validation_results.json

依赖：
    python -m pip install openpyxl
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter, defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel


# ============================================================
# 1. 三张表的结构与比较规则
# ============================================================

SCHEMAS: Dict[str, Dict[str, Any]] = {
    "subscription_flow": {
        "gold_sheet": "subscription_flow_gold",
        "extracted_sheet": "subscription_flow",

        "headers": [
            "公司",
            "事件类型",
            "时间",
            "认购人",
            "数量（万股）",
            "金额（万元）",
            "变化后总股本（万元）",
            "页码",
            "原文证据",
        ],

        # 不再将公司作为候选匹配键
        "block_fields": [
            "认购人",
        ],

        "core_fields": [
            "公司",
            "时间",
            "认购人",
            "数量（万股）",
            "金额（万元）",
            "变化后总股本（万元）",
        ],

        "soft_fields": [
            "事件类型",
            "页码",
            "原文证据",
        ],

        "required_fields": [
            "公司",
            "事件类型",
            "时间",
            "认购人",
        ],

        "numeric_fields": [
            "数量（万股）",
            "金额（万元）",
            "变化后总股本（万元）",
        ],
    },

    "share_transfer_flow": {
        "gold_sheet": "share_transfer_flow_gold",
        "extracted_sheet": "share_transfer_flow",

        "headers": [
            "公司",
            "事件类型",
            "时间",
            "转让方",
            "受让方",
            "数量（万元）",
            "价格",
            "页码",
            "原文证据",
        ],

        "block_fields": [
            "转让方",
            "受让方",
        ],

        "core_fields": [
            "公司",
            "时间",
            "转让方",
            "受让方",
            "数量（万元）",
        ],

        "soft_fields": [
            "事件类型",
            "价格",
            "页码",
            "原文证据",
        ],

        "required_fields": [
            "公司",
            "事件类型",
            "时间",
            "转让方",
            "受让方",
            "数量（万元）",
        ],

        "numeric_fields": [
            "数量（万元）",
        ],
    },

    "equity_snapshot": {
        "gold_sheet": "equity_snapshot_gold",
        "extracted_sheet": "equity_snapshot",

        "headers": [
            "公司",
            "时点",
            "总股本（万股）",
            "总出资额（万元）",
            "股东名称",
            "持股数（万股）",
            "出资额（万元）",
            "持股比例（%）",
            "页码",
            "原文证据",
        ],

        "block_fields": [
            "股东名称",
        ],

        "core_fields": [
            "公司",
            "时点",
            "总股本（万股）",
            "总出资额（万元）",
            "股东名称",
            "持股数（万股）",
            "出资额（万元）",
            "持股比例（%）",
        ],

        "soft_fields": [
            "页码",
            "原文证据",
        ],

        "required_fields": [
            "公司",
            "时点",
            "股东名称",
            "持股比例（%）",
        ],

        "numeric_fields": [
            "总股本（万股）",
            "总出资额（万元）",
            "持股数（万股）",
            "出资额（万元）",
            "持股比例（%）",
        ],
    },
}


ENTITY_FIELDS = {
    "公司",
    "认购人",
    "转让方",
    "受让方",
    "股东名称",
}

EVIDENCE_FIELDS = {
    "页码",
    "原文证据",
}


SANLIAN_COMPANY_ALIASES = {
    "三联锻造",
    "三联有限",
    "三联锻造有限公司",
    "三联锻造股份有限公司",
    "芜湖三联锻造有限公司",
    "芜湖三联锻造股份有限公司",
}


# ============================================================
# 2. 文本、主体与公司名称规范化
# ============================================================

def clean(value: Any) -> str:
    """去除换行和多余空格。"""

    if value is None:
        return ""

    text = str(value)

    text = (
        text
        .replace("\r", " ")
        .replace("\n", " ")
    )

    return re.sub(
        r"\s+",
        " ",
        text,
    ).strip()


def compact(value: Any) -> str:
    """删除全部空格并转换为小写。"""

    return re.sub(
        r"\s+",
        "",
        clean(value),
    ).lower()


def header_token(value: Any) -> str:
    """规范表头，用于字段匹配。"""

    return (
        compact(value)
        .replace("(", "（")
        .replace(")", "）")
        .replace("％", "%")
        .replace("﹪", "%")
        .replace("_", "")
    )


def normalize_entity(value: Any) -> str:
    """规范认购人、股东及转让主体名称。"""

    return (
        compact(value)
        .replace("（", "(")
        .replace("）", ")")
        .replace("，", ",")
        .replace("　", "")
    )


def normalize_company(value: Any) -> str:
    """
    将三联锻造历史名称和简称统一为“三联锻造”。
    """

    text = normalize_entity(value)

    alias_tokens = {
        normalize_entity(name)
        for name in SANLIAN_COMPANY_ALIASES
    }

    if text in alias_tokens:
        return "三联锻造"

    if (
        "三联锻造" in text
        or text == "三联有限"
    ):
        return "三联锻造"

    return text


def normalize_event_type(value: Any) -> str:
    """统一语义相同或相近的事件类型。"""

    text = compact(value)

    if not text:
        return ""

    if (
        "设立" in text
        or "初始出资" in text
    ):
        return "设立出资"

    if (
        "增资" in text
        or "新增股份" in text
        or "认购" in text
    ):
        return "增资"

    if (
        "股权转让" in text
        or "股份转让" in text
        or "老股转让" in text
    ):
        return "股权转让"

    if (
        "整体变更" in text
        or "股份制改造" in text
    ):
        return "整体变更"

    if "减资" in text:
        return "减资"

    if "转增" in text:
        return "转增"

    return text


# ============================================================
# 3. 数值规范化
# ============================================================

def parse_decimal(
    value: Any,
) -> Optional[Decimal]:
    """尝试将值转换为 Decimal。"""

    if (
        value is None
        or isinstance(value, bool)
    ):
        return None

    if isinstance(
        value,
        (int, float, Decimal),
    ):
        try:
            return Decimal(str(value))

        except InvalidOperation:
            return None

    text = clean(value)

    if not text:
        return None

    if text.lower() in {
        "null",
        "none",
        "na",
        "n/a",
        "未披露",
        "不适用",
        "无法确定",
    }:
        return None

    text = (
        text
        .replace(",", "")
        .replace("，", "")
        .replace("%", "")
        .replace("％", "")
    )

    text = re.sub(
        r"^(约|大约|合计|共)",
        "",
        text,
    )

    text = re.sub(
        r"(万元|万股|元|股)$",
        "",
        text,
    )

    if not re.fullmatch(
        r"[-+]?(?:\d+\.?\d*|\.\d+)",
        text,
    ):
        return None

    try:
        return Decimal(text)

    except InvalidOperation:
        return None


def normalize_number(value: Any) -> str:
    """
    将175、175.0、175.00统一为175。
    """

    number = parse_decimal(value)

    if number is None:
        return clean(value)

    text = format(
        number.normalize(),
        "f",
    )

    if text in {
        "",
        "-0",
    }:
        return "0"

    return text


# ============================================================
# 4. 时间规范化
# ============================================================

def normalize_month(
    value: Any,
    epoch: datetime,
) -> str:
    """
    将事件时间统一为 YYYY-MM。

    支持：
    - 2004年6月
    - 2004-06
    - 2004-06-18
    - Excel日期对象
    - Excel日期序列号
    """

    if value is None:
        return ""

    if isinstance(
        value,
        (datetime, date),
    ):
        return value.strftime(
            "%Y-%m"
        )

    # Excel数字日期序列号
    if (
        isinstance(value, (int, float))
        and not isinstance(value, bool)
    ):
        serial = float(value)

        if 20000 <= serial <= 80000:
            try:
                converted = from_excel(
                    serial,
                    epoch,
                )

                return converted.strftime(
                    "%Y-%m"
                )

            except Exception:
                pass

    text = clean(value)

    # 字符串形式的Excel日期序列号
    if re.fullmatch(
        r"\d+(?:\.0+)?",
        text,
    ):
        serial = float(text)

        if 20000 <= serial <= 80000:
            try:
                converted = from_excel(
                    serial,
                    epoch,
                )

                return converted.strftime(
                    "%Y-%m"
                )

            except Exception:
                pass

    # 中文年月
    match = re.search(
        r"(\d{4})年\s*0?(\d{1,2})月",
        text,
    )

    if match:
        return (
            f"{int(match.group(1)):04d}-"
            f"{int(match.group(2)):02d}"
        )

    # ISO年月或日期
    match = re.search(
        r"(\d{4})[-/.](\d{1,2})",
        text,
    )

    if match:
        return (
            f"{int(match.group(1)):04d}-"
            f"{int(match.group(2)):02d}"
        )

    if re.fullmatch(
        r"\d{4}",
        text,
    ):
        return text

    return compact(text)


# ============================================================
# 5. 时点规范化
# ============================================================

def normalize_timepoint(value: Any) -> str:
    """规范Pre-t0、t0、t1、t2等时点。"""

    text = clean(value)

    text = text.replace(
        "｜",
        "|",
    )

    text = re.sub(
        r"\s*\|\s*",
        "|",
        text,
    )

    text = re.sub(
        r"(?i)^pre[-_ ]?t0",
        "Pre-t0",
        text,
    )

    text = re.sub(
        r"(?i)^t(\d+)",
        lambda match: (
            f"t{int(match.group(1))}"
        ),
        text,
    )

    return text


def timepoint_key(value: Any) -> str:
    """
    生成用于Cross-check的时点键。

    t0与t0|报告期初视为同一时点；
    t1的不同描述视为同一时点；
    Pre-t0历史记录先归为同一类，再根据数值寻找最佳记录。
    """

    text = normalize_timepoint(value)

    match = re.match(
        r"(?i)^t(\d+)(?:\||$)",
        text,
    )

    if match:
        return (
            f"t{int(match.group(1))}"
        )

    if re.match(
        r"(?i)^pre-t0(?:\||$)",
        text,
    ):
        return "pre-t0"

    return compact(text)


# ============================================================
# 6. 页码与价格规范化
# ============================================================

def normalize_page(value: Any) -> str:
    """规范页码显示。"""

    text = clean(value)

    text = re.sub(
        r"原始\s*PDF\s*",
        "",
        text,
        flags=re.IGNORECASE,
    )

    for phrase in (
        "物理页码",
        "物理页",
        "页码",
        "第",
        "页",
    ):
        text = text.replace(
            phrase,
            "",
        )

    return (
        text
        .replace("，", "、")
        .replace(",", "、")
        .replace("至", "-")
        .replace("—", "-")
        .replace("–", "-")
        .replace(" ", "")
        .strip("、")
    )


def normalize_price(value: Any) -> str:
    """
    将1.00元/股和1元/股统一。
    """

    text = compact(value)

    return re.sub(
        r"\d+(?:\.\d+)?",
        lambda match: normalize_number(
            match.group(0)
        ),
        text,
    )


# ============================================================
# 7. 工作簿读取
# ============================================================

def header_lookup(
    table_name: str,
) -> Dict[str, str]:
    """生成实际表头到标准表头的映射。"""

    return {
        header_token(header): header
        for header
        in SCHEMAS[
            table_name
        ]["headers"]
    }


def find_header_row(
    worksheet,
    table_name: str,
) -> Tuple[int, List[str]]:
    """在前10行中自动寻找最可能的表头行。"""

    lookup = header_lookup(
        table_name
    )

    best_row = 0
    best_headers: List[str] = []
    best_score = -1

    for row_number in range(
        1,
        min(
            worksheet.max_row,
            10,
        ) + 1,
    ):
        headers = [
            clean(cell.value)
            for cell
            in worksheet[row_number]
        ]

        score = sum(
            1
            for header in headers
            if header_token(header)
            in lookup
        )

        if score > best_score:
            best_row = row_number
            best_headers = headers
            best_score = score

    minimum_score = max(
        3,
        len(
            SCHEMAS[
                table_name
            ]["headers"]
        ) // 2,
    )

    if best_score < minimum_score:
        raise ValueError(
            f"工作表 {worksheet.title} 无法识别有效表头；"
            f"识别字段数={best_score}，"
            f"至少需要={minimum_score}。"
        )

    while (
        best_headers
        and not best_headers[-1]
    ):
        best_headers.pop()

    return (
        best_row,
        best_headers,
    )


def read_table(
    workbook,
    table_name: str,
    sheet_name: str,
) -> Dict[str, Any]:
    """读取指定工作表。"""

    if sheet_name not in workbook.sheetnames:
        raise ValueError(
            f"工作簿缺少工作表：{sheet_name}；"
            f"实际工作表：{workbook.sheetnames}"
        )

    worksheet = workbook[
        sheet_name
    ]

    (
        header_row,
        raw_headers,
    ) = find_header_row(
        worksheet,
        table_name,
    )

    lookup = header_lookup(
        table_name
    )

    canonical_headers = [
        lookup.get(
            header_token(header),
            clean(header),
        )
        for header
        in raw_headers
    ]

    rows: List[
        Dict[str, Any]
    ] = []

    for row_number in range(
        header_row + 1,
        worksheet.max_row + 1,
    ):
        values = [
            worksheet.cell(
                row_number,
                column,
            ).value
            for column in range(
                1,
                len(raw_headers) + 1,
            )
        ]

        if not any(
            clean(value)
            for value in values
        ):
            continue

        record: Dict[
            str,
            Any,
        ] = {
            "_row": row_number
        }

        for index, field in enumerate(
            canonical_headers
        ):
            if field:
                record[field] = (
                    values[index]
                    if index < len(values)
                    else None
                )

        rows.append(record)

    return {
        "sheet_name": sheet_name,
        "headers": canonical_headers,
        "raw_headers": raw_headers,
        "rows": rows,
        "epoch": workbook.epoch,
    }


def read_inputs(
    gold_path: Path,
    extracted_path: Path,
) -> Tuple[
    Dict[str, Dict[str, Any]],
    Dict[str, Dict[str, Any]],
]:
    """读取Gold与抽取工作簿。"""

    if not gold_path.exists():
        raise FileNotFoundError(
            f"Gold文件不存在："
            f"{gold_path}"
        )

    if not extracted_path.exists():
        raise FileNotFoundError(
            f"抽取文件不存在："
            f"{extracted_path}"
        )

    gold_workbook = load_workbook(
        gold_path,
        data_only=True,
    )

    extracted_workbook = load_workbook(
        extracted_path,
        data_only=True,
    )

    gold_tables: Dict[
        str,
        Dict[str, Any],
    ] = {}

    extracted_tables: Dict[
        str,
        Dict[str, Any],
    ] = {}

    for (
        table_name,
        schema,
    ) in SCHEMAS.items():

        gold_tables[
            table_name
        ] = read_table(
            gold_workbook,
            table_name,
            schema["gold_sheet"],
        )

        extracted_tables[
            table_name
        ] = read_table(
            extracted_workbook,
            table_name,
            schema["extracted_sheet"],
        )

    return (
        gold_tables,
        extracted_tables,
    )


# ============================================================
# 8. 行规范化
# ============================================================

def normalize_row(
    table_name: str,
    row: Dict[str, Any],
    epoch: datetime,
) -> Dict[str, str]:
    """规范一条记录中的全部字段。"""

    schema = SCHEMAS[
        table_name
    ]

    result: Dict[
        str,
        str,
    ] = {}

    for field in schema[
        "headers"
    ]:
        value = row.get(field)

        if field in schema[
            "numeric_fields"
        ]:
            result[field] = (
                normalize_number(value)
            )

        elif field == "公司":
            result[field] = (
                normalize_company(value)
            )

        elif field in ENTITY_FIELDS:
            result[field] = (
                normalize_entity(value)
            )

        elif field == "事件类型":
            result[field] = (
                normalize_event_type(value)
            )

        elif field == "时间":
            result[field] = (
                normalize_month(
                    value,
                    epoch,
                )
            )

        elif field == "时点":
            result[field] = (
                normalize_timepoint(value)
            )

        elif field == "价格":
            result[field] = (
                normalize_price(value)
            )

        elif field == "页码":
            result[field] = (
                normalize_page(value)
            )

        elif field == "原文证据":
            result[field] = re.sub(
                r"\s+",
                "",
                clean(value),
            )

        else:
            result[field] = (
                compact(value)
            )

    return result


# ============================================================
# 9. 字段比较
# ============================================================

def comparable_value(
    table_name: str,
    field: str,
    value: str,
) -> str:
    """获取用于比较的字段值。"""

    if (
        table_name
        == "equity_snapshot"
        and field == "时点"
    ):
        return timepoint_key(
            value
        )

    return value


def values_equal(
    table_name: str,
    field: str,
    left: str,
    right: str,
    numeric_tolerance: Decimal,
) -> bool:
    """比较两个字段。"""

    if field in SCHEMAS[
        table_name
    ]["numeric_fields"]:

        left_number = parse_decimal(
            left
        )

        right_number = parse_decimal(
            right
        )

        if (
            left_number is None
            and right_number is None
        ):
            return (
                clean(left)
                == clean(right)
            )

        if (
            left_number is None
            or right_number is None
        ):
            return False

        return (
            abs(
                left_number
                - right_number
            )
            <= numeric_tolerance
        )

    return (
        comparable_value(
            table_name,
            field,
            left,
        )
        ==
        comparable_value(
            table_name,
            field,
            right,
        )
    )


def get_differences(
    table_name: str,
    gold_row: Dict[str, str],
    extracted_row: Dict[str, str],
    numeric_tolerance: Decimal,
) -> Tuple[
    List[str],
    List[str],
]:
    """返回核心差异和软差异字段。"""

    schema = SCHEMAS[
        table_name
    ]

    core_differences: List[
        str
    ] = []

    soft_differences: List[
        str
    ] = []

    for field in schema[
        "core_fields"
    ]:
        if not values_equal(
            table_name,
            field,
            gold_row.get(
                field,
                "",
            ),
            extracted_row.get(
                field,
                "",
            ),
            numeric_tolerance,
        ):
            core_differences.append(
                field
            )

    for field in schema[
        "soft_fields"
    ]:
        if not values_equal(
            table_name,
            field,
            gold_row.get(
                field,
                "",
            ),
            extracted_row.get(
                field,
                "",
            ),
            numeric_tolerance,
        ):
            soft_differences.append(
                field
            )

    return (
        core_differences,
        soft_differences,
    )


def block_key(
    table_name: str,
    row: Dict[str, str],
) -> Tuple[str, ...]:
    """
    生成候选配对键。

    注意：三个表均已移除公司字段。
    """

    return tuple(
        row.get(
            field,
            "",
        )
        for field
        in SCHEMAS[
            table_name
        ]["block_fields"]
    )


def candidate_score(
    table_name: str,
    core_differences: List[str],
    soft_differences: List[str],
) -> float:
    """
    同一主体存在多条记录时，
    选择核心差异最少的候选。
    """

    score = (
        len(
            soft_differences
        )
        * 0.1
    )

    numeric_fields = set(
        SCHEMAS[
            table_name
        ]["numeric_fields"]
    )

    for field in core_differences:

        if field in {
            "时间",
            "时点",
        }:
            score += 10

        elif field in numeric_fields:
            score += 6

        elif field == "公司":
            score += 4

        else:
            score += 3

    return score


# ============================================================
# 10. Cross-check
# ============================================================

def build_detail(
    table_name: str,
    status: str,
    gold_raw: Optional[
        Dict[str, Any]
    ],
    extracted_raw: Optional[
        Dict[str, Any]
    ],
    key: Tuple[str, ...],
    core_differences: List[str],
    soft_differences: List[str],
) -> Dict[str, Any]:
    """生成Cross-check明细记录。"""

    row: Dict[
        str,
        Any,
    ] = {
        "表名": table_name,

        "状态": status,

        "Gold行号": (
            ""
            if gold_raw is None
            else gold_raw.get(
                "_row",
                "",
            )
        ),

        "抽取表行号": (
            ""
            if extracted_raw is None
            else extracted_raw.get(
                "_row",
                "",
            )
        ),

        "主体匹配键": (
            " | ".join(key)
        ),

        "核心差异字段": (
            "、".join(
                core_differences
            )
        ),

        "软差异字段": (
            "、".join(
                soft_differences
            )
        ),
    }

    for field in SCHEMAS[
        table_name
    ]["headers"]:

        row[
            f"Gold_{field}"
        ] = (
            ""
            if gold_raw is None
            else gold_raw.get(
                field,
                "",
            )
        )

        row[
            f"抽取_{field}"
        ] = (
            ""
            if extracted_raw is None
            else extracted_raw.get(
                field,
                "",
            )
        )

    return row


def cross_check_table(
    table_name: str,
    gold_data: Dict[str, Any],
    extracted_data: Dict[str, Any],
    numeric_tolerance: Decimal,
) -> Tuple[
    Dict[str, Any],
    List[Dict[str, Any]],
]:
    """对一张表执行Cross-check。"""

    gold_raw = gold_data[
        "rows"
    ]

    extracted_raw = extracted_data[
        "rows"
    ]

    gold_rows = [
        normalize_row(
            table_name,
            row,
            gold_data["epoch"],
        )
        for row in gold_raw
    ]

    extracted_rows = [
        normalize_row(
            table_name,
            row,
            extracted_data["epoch"],
        )
        for row in extracted_raw
    ]

    buckets: Dict[
        Tuple[str, ...],
        List[int],
    ] = defaultdict(list)

    for index, row in enumerate(
        extracted_rows
    ):
        buckets[
            block_key(
                table_name,
                row,
            )
        ].append(index)

    used: set[int] = set()

    details: List[
        Dict[str, Any]
    ] = []

    exact_count = 0
    soft_only_count = 0
    core_difference_count = 0
    missing_count = 0

    for (
        gold_index,
        gold_row,
    ) in enumerate(
        gold_rows
    ):
        key = block_key(
            table_name,
            gold_row,
        )

        candidates = [
            index
            for index
            in buckets.get(
                key,
                [],
            )
            if index not in used
        ]

        if not candidates:
            missing_count += 1

            details.append(
                build_detail(
                    table_name,
                    "MISSING_IN_EXTRACTED",
                    gold_raw[
                        gold_index
                    ],
                    None,
                    key,
                    [],
                    [],
                )
            )

            continue

        ranked = []

        for candidate_index in candidates:
            (
                core_differences,
                soft_differences,
            ) = get_differences(
                table_name,
                gold_row,
                extracted_rows[
                    candidate_index
                ],
                numeric_tolerance,
            )

            ranked.append(
                (
                    candidate_score(
                        table_name,
                        core_differences,
                        soft_differences,
                    ),
                    candidate_index,
                    core_differences,
                    soft_differences,
                )
            )

        (
            _,
            best_index,
            core_differences,
            soft_differences,
        ) = min(
            ranked,
            key=lambda item: (
                item[0],
                item[1],
            ),
        )

        used.add(
            best_index
        )

        if (
            not core_differences
            and not soft_differences
        ):
            status = "EXACT_MATCH"
            exact_count += 1

        elif not core_differences:
            status = (
                "MATCH_WITH_SOFT_DIFFERENCES"
            )
            soft_only_count += 1

        else:
            status = (
                "MATCH_WITH_CORE_DIFFERENCES"
            )
            core_difference_count += 1

        details.append(
            build_detail(
                table_name,
                status,
                gold_raw[
                    gold_index
                ],
                extracted_raw[
                    best_index
                ],
                key,
                core_differences,
                soft_differences,
            )
        )

    # 抽取表多余记录只提示，不阻断PASS
    extra_indices = [
        index
        for index in range(
            len(
                extracted_rows
            )
        )
        if index not in used
    ]

    for index in extra_indices:
        details.append(
            build_detail(
                table_name,
                "EXTRA_IN_EXTRACTED_WARNING",
                None,
                extracted_raw[
                    index
                ],
                block_key(
                    table_name,
                    extracted_rows[
                        index
                    ],
                ),
                [],
                [],
            )
        )

    coverage = (
        1.0
        if not gold_rows
        else (
            len(gold_rows)
            - missing_count
        )
        / len(gold_rows)
    )

    table_pass = (
        missing_count == 0
        and core_difference_count == 0
    )

    summary = {
        "表名": table_name,

        "Gold行数": len(
            gold_rows
        ),

        "抽取表行数": len(
            extracted_rows
        ),

        "完全一致": (
            exact_count
        ),

        "仅软字段不同": (
            soft_only_count
        ),

        "核心字段不同": (
            core_difference_count
        ),

        "抽取缺失": (
            missing_count
        ),

        "抽取多余（仅警告）": len(
            extra_indices
        ),

        "Gold覆盖率": round(
            coverage,
            6,
        ),

        "通过": table_pass,
    }

    return (
        summary,
        details,
    )


# ============================================================
# 11. Schema校验
# ============================================================

def schema_issue(
    table_name: str,
    level: str,
    check: str,
    row_number: Any,
    field: str,
    message: str,
) -> Dict[str, Any]:
    """创建Schema问题记录。"""

    return {
        "来源": "EXTRACTED",
        "表名": table_name,
        "级别": level,
        "检查项": check,
        "行号": row_number,
        "字段": field,
        "说明": message,
    }


def validate_schema(
    table_name: str,
    table_data: Dict[str, Any],
    numeric_tolerance: Decimal,
    ratio_tolerance: Decimal,
) -> List[Dict[str, Any]]:
    """
    只对抽取工作簿进行Schema校验。
    """

    schema = SCHEMAS[
        table_name
    ]

    issues: List[
        Dict[str, Any]
    ] = []

    actual_headers = [
        header
        for header
        in table_data[
            "headers"
        ]
        if header
    ]

    missing_headers = [
        header
        for header
        in schema[
            "headers"
        ]
        if header not in actual_headers
    ]

    if missing_headers:
        issues.append(
            schema_issue(
                table_name,
                "ERROR",
                "表头完整性",
                1,
                "",
                (
                    "缺少字段："
                    + "、".join(
                        missing_headers
                    )
                ),
            )
        )

    normalized_rows = [
        normalize_row(
            table_name,
            row,
            table_data["epoch"],
        )
        for row
        in table_data[
            "rows"
        ]
    ]

    for index, row in enumerate(
        normalized_rows
    ):
        row_number = (
            table_data[
                "rows"
            ][index].get(
                "_row",
                index + 2,
            )
        )

        # 核心必填字段
        for field in schema[
            "required_fields"
        ]:
            if not row.get(
                field,
                "",
            ):
                issues.append(
                    schema_issue(
                        table_name,
                        "ERROR",
                        "核心必填字段",
                        row_number,
                        field,
                        "字段为空。",
                    )
                )

        # 数值格式
        for field in schema[
            "numeric_fields"
        ]:
            value = row.get(
                field,
                "",
            )

            if (
                value
                and parse_decimal(
                    value
                ) is None
            ):
                issues.append(
                    schema_issue(
                        table_name,
                        "ERROR",
                        "数值类型",
                        row_number,
                        field,
                        (
                            "无法解析为数值："
                            f"{value}"
                        ),
                    )
                )

        # 页码和证据为空只记WARNING
        for field in EVIDENCE_FIELDS:
            if (
                field
                in schema[
                    "headers"
                ]
                and not row.get(
                    field,
                    "",
                )
            ):
                issues.append(
                    schema_issue(
                        table_name,
                        "WARNING",
                        "证据字段",
                        row_number,
                        field,
                        (
                            "字段为空，"
                            "建议人工复核。"
                        ),
                    )
                )

    # 完全重复记录只记WARNING
    signatures = [
        tuple(
            row.get(
                header,
                "",
            )
            for header
            in schema[
                "headers"
            ]
        )
        for row
        in normalized_rows
    ]

    for (
        _,
        count,
    ) in Counter(
        signatures
    ).items():
        if count > 1:
            issues.append(
                schema_issue(
                    table_name,
                    "WARNING",
                    "完全重复记录",
                    "",
                    "",
                    (
                        f"发现{count}条"
                        "完全重复记录。"
                    ),
                )
            )

    # equity_snapshot闭环问题记WARNING
    if (
        table_name
        == "equity_snapshot"
    ):
        groups: Dict[
            Tuple[str, str],
            List[Dict[str, str]],
        ] = defaultdict(list)

        for row in normalized_rows:
            groups[
                (
                    row.get(
                        "公司",
                        "",
                    ),
                    timepoint_key(
                        row.get(
                            "时点",
                            "",
                        )
                    ),
                )
            ].append(row)

        for (
            company,
            point,
        ), group in groups.items():

            label = (
                f"{company}|{point}"
            )

            names = [
                row.get(
                    "股东名称",
                    "",
                )
                for row in group
            ]

            duplicates = [
                name
                for (
                    name,
                    count,
                ) in Counter(
                    names
                ).items()
                if (
                    name
                    and count > 1
                )
            ]

            if duplicates:
                issues.append(
                    schema_issue(
                        table_name,
                        "WARNING",
                        "一行一股东",
                        "",
                        "股东名称",
                        (
                            f"{label}存在重复股东："
                            f"{'、'.join(duplicates)}"
                        ),
                    )
                )

            total_shares = {
                parse_decimal(
                    row.get(
                        "总股本（万股）"
                    )
                )
                for row in group
            }

            total_shares.discard(
                None
            )

            holding_values = [
                parse_decimal(
                    row.get(
                        "持股数（万股）"
                    )
                )
                for row in group
            ]

            if (
                len(total_shares) == 1
                and all(
                    value is not None
                    for value
                    in holding_values
                )
            ):
                stated = next(
                    iter(
                        total_shares
                    )
                )

                calculated = sum(
                    value
                    for value
                    in holding_values
                    if value is not None
                )

                if (
                    abs(
                        calculated
                        - stated
                    )
                    > numeric_tolerance
                ):
                    issues.append(
                        schema_issue(
                            table_name,
                            "WARNING",
                            "持股数闭环",
                            "",
                            "持股数（万股）",
                            (
                                f"{label}持股数合计"
                                f"{calculated}，"
                                f"总股本{stated}。"
                            ),
                        )
                    )

            ratios = [
                parse_decimal(
                    row.get(
                        "持股比例（%）"
                    )
                )
                for row in group
            ]

            if (
                ratios
                and all(
                    value is not None
                    for value
                    in ratios
                )
            ):
                ratio_sum = sum(
                    value
                    for value
                    in ratios
                    if value is not None
                )

                if (
                    abs(
                        ratio_sum
                        - Decimal("100")
                    )
                    > ratio_tolerance
                ):
                    issues.append(
                        schema_issue(
                            table_name,
                            "WARNING",
                            "持股比例闭环",
                            "",
                            "持股比例（%）",
                            (
                                f"{label}持股比例合计"
                                f"{ratio_sum}。"
                            ),
                        )
                    )

    return issues


# ============================================================
# 12. 输出校验结果
# ============================================================

def append_rows(
    worksheet,
    headers: Sequence[str],
    rows: Iterable[
        Dict[str, Any]
    ],
) -> None:
    """向Excel工作表写入数据。"""

    worksheet.append(
        list(headers)
    )

    for row in rows:
        values: List[Any] = []

        for header in headers:
            value = row.get(
                header,
                "",
            )

            if isinstance(
                value,
                (dict, list),
            ):
                value = json.dumps(
                    value,
                    ensure_ascii=False,
                )

            values.append(value)

        worksheet.append(values)


def style_sheet(
    worksheet,
) -> None:
    """设置结果工作表格式。"""

    fill = PatternFill(
        "solid",
        fgColor="D9E2F3",
    )

    for cell in worksheet[1]:
        cell.fill = fill

        cell.font = Font(
            bold=True
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    worksheet.freeze_panes = "A2"

    worksheet.auto_filter.ref = (
        worksheet.dimensions
    )

    for (
        column_index,
        column,
    ) in enumerate(
        worksheet.iter_cols(),
        start=1,
    ):
        width = max(
            (
                len(str(cell.value))
                if cell.value is not None
                else 0
            )
            for cell in column
        )

        worksheet.column_dimensions[
            get_column_letter(
                column_index
            )
        ].width = min(
            max(
                width + 2,
                10,
            ),
            45,
        )

    for row in worksheet.iter_rows(
        min_row=2
    ):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )


def write_results(
    output_dir: Path,
    summaries: List[
        Dict[str, Any]
    ],
    details: List[
        Dict[str, Any]
    ],
    issues: List[
        Dict[str, Any]
    ],
    final_result: Dict[str, Any],
) -> None:
    """输出validation_results.xlsx和JSON。"""

    output_dir.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook = Workbook()

    workbook.remove(
        workbook.active
    )

    # Summary
    summary_sheet = workbook.create_sheet(
        "Summary"
    )

    append_rows(
        summary_sheet,
        [
            "指标",
            "值",
        ],
        [
            {
                "指标": "总体结果",
                "值": (
                    "PASS"
                    if final_result[
                        "overall_pass"
                    ]
                    else "FAIL / NEED REVIEW"
                ),
            },
            {
                "指标": (
                    "Cross-check是否通过"
                ),
                "值": final_result[
                    "crosscheck_pass"
                ],
            },
            {
                "指标": (
                    "Schema是否通过"
                ),
                "值": final_result[
                    "schema_pass"
                ],
            },
            {
                "指标": (
                    "Schema ERROR数量"
                ),
                "值": final_result[
                    "schema_error_count"
                ],
            },
            {
                "指标": (
                    "Schema WARNING数量"
                ),
                "值": final_result[
                    "schema_warning_count"
                ],
            },
            {
                "指标": "匹配改进",
                "值": (
                    "统一三联公司别名；"
                    "主体匹配键不含公司；"
                    "多余记录仅警告"
                ),
            },
            {
                "指标": "Gold文件",
                "值": final_result[
                    "inputs"
                ]["gold"],
            },
            {
                "指标": "抽取文件",
                "值": final_result[
                    "inputs"
                ]["extracted"],
            },
        ],
    )

    style_sheet(
        summary_sheet
    )

    # CrossCheck_Summary
    cross_summary_sheet = (
        workbook.create_sheet(
            "CrossCheck_Summary"
        )
    )

    append_rows(
        cross_summary_sheet,
        [
            "表名",
            "Gold行数",
            "抽取表行数",
            "完全一致",
            "仅软字段不同",
            "核心字段不同",
            "抽取缺失",
            "抽取多余（仅警告）",
            "Gold覆盖率",
            "通过",
        ],
        summaries,
    )

    style_sheet(
        cross_summary_sheet
    )

    # CrossCheck_Detail
    all_fields: List[str] = []

    for table_name in SCHEMAS:
        for field in SCHEMAS[
            table_name
        ]["headers"]:

            if field not in all_fields:
                all_fields.append(
                    field
                )

    detail_sheet = workbook.create_sheet(
        "CrossCheck_Detail"
    )

    detail_headers = [
        "表名",
        "状态",
        "Gold行号",
        "抽取表行号",
        "主体匹配键",
        "核心差异字段",
        "软差异字段",
    ]

    detail_headers += [
        f"Gold_{field}"
        for field in all_fields
    ]

    detail_headers += [
        f"抽取_{field}"
        for field in all_fields
    ]

    append_rows(
        detail_sheet,
        detail_headers,
        details,
    )

    style_sheet(
        detail_sheet
    )

    # Schema_Issues
    schema_sheet = workbook.create_sheet(
        "Schema_Issues"
    )

    append_rows(
        schema_sheet,
        [
            "来源",
            "表名",
            "级别",
            "检查项",
            "行号",
            "字段",
            "说明",
        ],
        issues,
    )

    style_sheet(
        schema_sheet
    )

    workbook.save(
        output_dir
        / "validation_results.xlsx"
    )

    (
        output_dir
        / "validation_results.json"
    ).write_text(
        json.dumps(
            final_result,
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )


# ============================================================
# 13. 主程序
# ============================================================

def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "sanlian_gold.xlsx与sanlian.xlsx"
            "的Cross-check和Schema校验"
        )
    )

    parser.add_argument(
        "--gold",
        required=True,
        help="sanlian_gold.xlsx路径",
    )

    parser.add_argument(
        "--extracted",
        required=True,
        help="sanlian.xlsx路径",
    )

    parser.add_argument(
        "--output-dir",
        required=True,
        help="校验结果输出目录",
    )

    parser.add_argument(
        "--numeric-tolerance",
        default="0.01",
        help=(
            "数值字段允许的绝对误差，"
            "默认0.01"
        ),
    )

    parser.add_argument(
        "--ratio-tolerance",
        default="0.10",
        help=(
            "持股比例合计允许误差，"
            "默认0.10"
        ),
    )

    args = parser.parse_args()

    try:
        numeric_tolerance = Decimal(
            args.numeric_tolerance
        )

        ratio_tolerance = Decimal(
            args.ratio_tolerance
        )

    except InvalidOperation as error:
        raise ValueError(
            "容差参数必须是合法数字。"
        ) from error

    gold_path = Path(
        args.gold
    )

    extracted_path = Path(
        args.extracted
    )

    output_dir = Path(
        args.output_dir
    )

    (
        gold_tables,
        extracted_tables,
    ) = read_inputs(
        gold_path,
        extracted_path,
    )

    summaries: List[
        Dict[str, Any]
    ] = []

    details: List[
        Dict[str, Any]
    ] = []

    issues: List[
        Dict[str, Any]
    ] = []

    for table_name in SCHEMAS:
        (
            summary,
            table_details,
        ) = cross_check_table(
            table_name,
            gold_tables[
                table_name
            ],
            extracted_tables[
                table_name
            ],
            numeric_tolerance,
        )

        summaries.append(
            summary
        )

        details.extend(
            table_details
        )

        issues.extend(
            validate_schema(
                table_name,
                extracted_tables[
                    table_name
                ],
                numeric_tolerance,
                ratio_tolerance,
            )
        )

    error_count = sum(
        1
        for item in issues
        if item["级别"] == "ERROR"
    )

    warning_count = sum(
        1
        for item in issues
        if item["级别"]
        == "WARNING"
    )

    crosscheck_pass = all(
        item["通过"]
        for item in summaries
    )

    schema_pass = (
        error_count == 0
    )

    overall_pass = (
        crosscheck_pass
        and schema_pass
    )

    final_result = {
        "overall_pass": (
            overall_pass
        ),

        "crosscheck_pass": (
            crosscheck_pass
        ),

        "schema_pass": (
            schema_pass
        ),

        "schema_error_count": (
            error_count
        ),

        "schema_warning_count": (
            warning_count
        ),

        "inputs": {
            "gold": str(
                gold_path.resolve()
            ),

            "extracted": str(
                extracted_path.resolve()
            ),
        },

        "sheet_mapping": {
            table_name: {
                "gold": (
                    SCHEMAS[
                        table_name
                    ]["gold_sheet"]
                ),

                "extracted": (
                    SCHEMAS[
                        table_name
                    ]["extracted_sheet"]
                ),
            }
            for table_name
            in SCHEMAS
        },

        "company_aliases": sorted(
            SANLIAN_COMPANY_ALIASES
        ),

        "crosscheck": summaries,

        "schema_issues": issues,
    }

    write_results(
        output_dir,
        summaries,
        details,
        issues,
        final_result,
    )

    print("=" * 78)

    print(
        f"Gold："
        f"{gold_path.resolve()}"
    )

    print(
        f"抽取表："
        f"{extracted_path.resolve()}"
    )

    for item in summaries:
        print(
            f"{item['表名']}: "
            f"Gold={item['Gold行数']}, "
            f"抽取={item['抽取表行数']}, "
            f"完全一致={item['完全一致']}, "
            f"软差异={item['仅软字段不同']}, "
            f"核心差异={item['核心字段不同']}, "
            f"缺失={item['抽取缺失']}, "
            f"多余={item['抽取多余（仅警告）']}, "
            f"通过={item['通过']}"
        )

    print(
        f"Schema ERROR={error_count}, "
        f"WARNING={warning_count}, "
        f"通过={schema_pass}"
    )

    print(
        "最终结果：{}".format(
            "PASS"
            if overall_pass
            else "FAIL / NEED REVIEW"
        )
    )

    print(
        "结果文件：{}".format(
            (
                output_dir
                / "validation_results.xlsx"
            ).resolve()
        )
    )

    print("=" * 78)


if __name__ == "__main__":
    try:
        main()

    except Exception as error:
        print(
            f"校验失败：{error}",
            file=sys.stderr,
        )

        sys.exit(1)