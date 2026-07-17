from __future__ import annotations
from pathlib import Path
from typing import Dict, Any, List
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SHEETS = {
    "company": "公司信息",
    "capital_increase": "认购增资",
    "equity_transfer": "股权转让",
    "equity_snapshot": "股权快照",
}

def _records(value):
    if isinstance(value, list):
        return value
    if isinstance(value, dict):
        return [value]
    return []

def export_workbook(data: Dict[str, Any], issues: List[Dict[str, Any]],
                    output_path: Path, crosscheck: List[Dict[str, Any]] | None = None):
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for key, sheet_name in SHEETS.items():
            pd.DataFrame(_records(data.get(key, {}))).to_excel(
                writer, sheet_name=sheet_name, index=False
            )
        pd.DataFrame(issues).to_excel(writer, sheet_name="校验问题", index=False)
        if crosscheck is not None:
            pd.DataFrame(crosscheck).to_excel(writer, sheet_name="Gold对比", index=False)

        wb = writer.book
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="D9EAF7")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for col_idx, column in enumerate(ws.columns, start=1):
                max_len = 0
                for cell in column:
                    value = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, min(len(value), 80))
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 45)

def export_batch_summary(rows: List[Dict[str, Any]], output_path: Path):
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        pd.DataFrame(rows).to_excel(writer, sheet_name="批处理汇总", index=False)
        ws = writer.book["批处理汇总"]
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
        for i, col in enumerate(ws.columns, start=1):
            width = max((len(str(c.value or "")) for c in col), default=10) + 2
            ws.column_dimensions[get_column_letter(i)].width = min(max(width, 10), 40)
